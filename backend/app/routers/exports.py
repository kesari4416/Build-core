from datetime import date
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from app.database import get_db
from app.models import User
from app.core.security import get_current_user
from app.routers.finance import FIN, STAFF, org_balance_sheet, project_balance_sheet

router = APIRouter(tags=["exports"])

ORANGE = colors.HexColor("#f97316")
DARK = colors.HexColor("#18181b")
RED = colors.HexColor("#dc2626")
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def money(n):
    return f"Rs {float(n or 0):,.2f}"


def _styles():
    ss = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("t", parent=ss["Title"], fontName="Helvetica-Bold",
                                fontSize=18, textColor=DARK, alignment=0, spaceAfter=2),
        "sub": ParagraphStyle("s", parent=ss["Normal"], fontSize=9,
                              textColor=colors.HexColor("#71717a"), spaceAfter=12),
        "h2": ParagraphStyle("h", parent=ss["Heading2"], fontName="Helvetica-Bold",
                             fontSize=11, textColor=ORANGE, spaceBefore=14, spaceAfter=6),
        "cell": ParagraphStyle("c", parent=ss["Normal"], fontSize=8),
    }


def styled_table(data, col_widths=None, red_rows=(), bold_last=False, left_cols=1):
    t = Table(data, colWidths=col_widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), DARK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d4d4d8")),
        ("ALIGN", (left_cols, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fafafa")]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    for r in red_rows:
        style.append(("TEXTCOLOR", (0, r), (-1, r), RED))
    if bold_last:
        style += [("FONTNAME", (0, len(data) - 1), (-1, len(data) - 1), "Helvetica-Bold"),
                  ("BACKGROUND", (0, len(data) - 1), (-1, len(data) - 1), colors.HexColor("#f4f4f5"))]
    t.setStyle(TableStyle(style))
    return t


def pdf_response(story, filename):
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    doc.build(story)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


def xlsx_response(wb, filename):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type=XLSX_MIME,
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


HEAD_FILL = PatternFill(start_color="18181B", end_color="18181B", fill_type="solid")
HEAD_FONT = Font(color="FFFFFF", bold=True)
RED_FONT = Font(color="DC2626", bold=True)
TITLE_FONT = Font(bold=True, size=14)
BOLD = Font(bold=True)


def ws_header_row(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        ws.column_dimensions[c.column_letter] = ws.column_dimensions[c.column_letter]
        ws.column_dimensions[c.column_letter].width = w


@router.get("/finance/balance-sheet/export")
def export_org_balance_sheet(fmt: Literal["pdf", "xlsx"] = "pdf",
                             db: Session = Depends(get_db), user: User = Depends(FIN)):
    bs = org_balance_sheet(db=db, user=user)
    today = date.today().isoformat()
    dues = bs["employee_dues"]
    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"
        ws["A1"] = "SITERA — All Projects Balance Sheet"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = f"Generated {today}"
        for i, (k, v) in enumerate([("Total Credit (In)", bs["total_credit"]),
                                    ("Total Debit (Out)", bs["total_debit"]),
                                    ("Overall Profit", bs["overall_profit"]),
                                    ("Overall Loss", bs["overall_loss"]),
                                    ("Net", bs["net"])], start=4):
            ws.cell(row=i, column=1, value=k).font = BOLD
            ws.cell(row=i, column=2, value=v).number_format = "#,##0.00"
        r0 = 10
        ws_header_row(ws, r0, ["Project", "Budget", "Credit (In)", "Debit (Out)", "Profit / Loss"],
                      [30, 16, 16, 16, 16])
        for j, p in enumerate(bs["projects"], start=1):
            row = [p["name"], p["budget"], p["credit"], p["debit"], p["profit_loss"]]
            for i, v in enumerate(row, start=1):
                c = ws.cell(row=r0 + j, column=i, value=v)
                if i > 1:
                    c.number_format = "#,##0.00"
                if p["is_loss"]:
                    c.font = RED_FONT
        tr = r0 + len(bs["projects"]) + 1
        for i, v in enumerate(["TOTAL", "", bs["total_credit"], bs["total_debit"], bs["net"]], start=1):
            c = ws.cell(row=tr, column=i, value=v)
            c.font = BOLD
            if i > 2:
                c.number_format = "#,##0.00"
        ws2 = wb.create_sheet("Employee Dues")
        ws2["A1"] = "Required Employee Payments"
        ws2["A1"].font = TITLE_FONT
        ws2["A3"] = "Staff payroll pending"
        ws2["B3"] = dues["staff_payroll_pending"]
        ws_header_row(ws2, 5, ["Labour category", "Amount due"], [26, 16])
        for j, c in enumerate(dues["labour_by_category"], start=1):
            ws2.cell(row=5 + j, column=1, value=c["category"])
            ws2.cell(row=5 + j, column=2, value=c["amount"]).number_format = "#,##0.00"
        tr2 = 6 + len(dues["labour_by_category"])
        ws2.cell(row=tr2, column=1, value="TOTAL REQUIRED").font = BOLD
        tc = ws2.cell(row=tr2, column=2, value=dues["total_required"])
        tc.font = BOLD
        tc.number_format = "#,##0.00"
        for col, w in (("A", 26), ("B", 16)):
            ws2.column_dimensions[col].width = w
        return xlsx_response(wb, f"sitera-balance-sheet-{today}.xlsx")

    st = _styles()
    story = [Paragraph("SITERA — All Projects Balance Sheet", st["title"]),
             Paragraph(f"Generated {today}", st["sub"])]
    story.append(styled_table(
        [["Total Credit (In)", "Total Debit (Out)", "Overall Profit", "Overall Loss", "Net"],
         [money(bs["total_credit"]), money(bs["total_debit"]), money(bs["overall_profit"]),
          money(bs["overall_loss"]), money(bs["net"])]], col_widths=[36 * mm] * 5, left_cols=0))
    story.append(Paragraph("Projects — Credit / Debit / Profit-Loss", st["h2"]))
    rows = [["Project", "Budget", "Credit (In)", "Debit (Out)", "Profit / Loss"]]
    red = []
    for i, p in enumerate(bs["projects"], start=1):
        rows.append([p["name"], money(p["budget"]), money(p["credit"]),
                     money(p["debit"]), money(p["profit_loss"])])
        if p["is_loss"]:
            red.append(i)
    rows.append(["TOTAL", "", money(bs["total_credit"]), money(bs["total_debit"]), money(bs["net"])])
    story.append(styled_table(rows, col_widths=[52 * mm, 32 * mm, 32 * mm, 32 * mm, 32 * mm],
                              red_rows=red, bold_last=True))
    if bs["loss_projects"]:
        story.append(Paragraph("Loss-Making Projects", st["h2"]))
        lrows = [["Project", "Loss"]] + [[lp["name"], money(lp["loss"])] for lp in bs["loss_projects"]]
        story.append(styled_table(lrows, col_widths=[110 * mm, 40 * mm], red_rows=range(1, len(lrows))))
    story.append(Paragraph("Required Employee Payments", st["h2"]))
    drows = [["Type", "Amount due"], ["Staff payroll pending", money(dues["staff_payroll_pending"])]]
    drows += [[f"Labour — {c['category']}", money(c["amount"])] for c in dues["labour_by_category"]]
    drows.append(["TOTAL REQUIRED", money(dues["total_required"])])
    story.append(styled_table(drows, col_widths=[110 * mm, 40 * mm], bold_last=True))
    return pdf_response(story, f"sitera-balance-sheet-{today}.pdf")


@router.get("/projects/{project_id}/balance-sheet/export")
def export_project_balance_sheet(project_id: int, fmt: Literal["pdf", "xlsx"] = "pdf",
                                 db: Session = Depends(get_db), user: User = Depends(STAFF)):
    bs = project_balance_sheet(project_id=project_id, db=db, user=user)
    today = date.today().isoformat()
    slug = "".join(ch if ch.isalnum() else "-" for ch in bs["name"].lower()).strip("-")
    r = bs["released"]
    summary = [("Total Budget", bs["budget"]), ("Client Payment (In)", bs["client_paid"]),
               ("Outstanding from Client", bs["client_outstanding"]),
               ("Payment Released (Out)", bs["total_released"]),
               ("Balance (In - Out)", bs["balance"]), ("Budget Remaining", bs["budget_remaining"])]
    breakdown = [("Daily-Wage Labour", r["labour_wages"]), ("Staff Payroll", r["staff_payroll"]),
                 ("Site Expenses", r["expenses"]), ("Procurement / Vendors", r["procurement"])]
    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws["A1"] = f"SITERA — Balance Sheet: {bs['name']}"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = f"Generated {today}"
        for i, (k, v) in enumerate(summary, start=4):
            ws.cell(row=i, column=1, value=k).font = BOLD
            ws.cell(row=i, column=2, value=v).number_format = "#,##0.00"
        ws.cell(row=11, column=1, value="Payment Released Breakdown").font = BOLD
        for i, (k, v) in enumerate(breakdown, start=12):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=v).number_format = "#,##0.00"
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18
        ws2 = wb.create_sheet("Transactions")
        ws_header_row(ws2, 1, ["Date", "Description", "Credit (In)", "Debit (Out)"], [13, 55, 16, 16])
        for j, en in enumerate(bs["entries"], start=2):
            ws2.cell(row=j, column=1, value=en["date"] or "")
            ws2.cell(row=j, column=2, value=en["description"])
            col = 3 if en["type"] == "credit" else 4
            ws2.cell(row=j, column=col, value=en["amount"]).number_format = "#,##0.00"
        tr = len(bs["entries"]) + 2
        ws2.cell(row=tr, column=2, value="TOTAL").font = BOLD
        for col, v in ((3, bs["total_credit"]), (4, bs["total_debit"])):
            c = ws2.cell(row=tr, column=col, value=v)
            c.font = BOLD
            c.number_format = "#,##0.00"
        return xlsx_response(wb, f"{slug}-balance-sheet-{today}.xlsx")

    st = _styles()
    story = [Paragraph(f"SITERA — Balance Sheet: {bs['name']}", st["title"]),
             Paragraph(f"Generated {today}", st["sub"])]
    story.append(styled_table([[k for k, _ in summary], [money(v) for _, v in summary]],
                              col_widths=[30 * mm] * 6, left_cols=0))
    story.append(Paragraph("Payment Released Breakdown — where the money is going", st["h2"]))
    story.append(styled_table([["Category", "Amount"]] + [[k, money(v)] for k, v in breakdown],
                              col_widths=[110 * mm, 40 * mm]))
    story.append(Paragraph("All Transactions — Latest First", st["h2"]))
    trows = [["Date", "Description", "Credit (In)", "Debit (Out)"]]
    for en in bs["entries"]:
        trows.append([en["date"] or "-", Paragraph(en["description"], st["cell"]),
                      money(en["amount"]) if en["type"] == "credit" else "",
                      money(en["amount"]) if en["type"] == "debit" else ""])
    trows.append(["", "TOTAL", money(bs["total_credit"]), money(bs["total_debit"])])
    story.append(styled_table(trows, col_widths=[22 * mm, 88 * mm, 35 * mm, 35 * mm],
                              bold_last=True, left_cols=2))
    return pdf_response(story, f"{slug}-balance-sheet-{today}.pdf")


@router.get("/projects/{project_id}/change-orders/export")
def export_change_orders(project_id: int, fmt: Literal["pdf", "xlsx"] = "pdf",
                         phase_id: int = None, category: str = None, status: str = None,
                         db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    from app.models import Phase, Project, ProjectChangeOrder
    from app.routers.change_orders import check_co_access, co_totals, f as cf

    project = db.get(Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    check_co_access(user, project)
    q = db.query(ProjectChangeOrder).filter(ProjectChangeOrder.project_id == project_id)
    if phase_id:
        q = q.filter(ProjectChangeOrder.phase_id == phase_id)
    if category:
        q = q.filter(ProjectChangeOrder.category == category)
    if status:
        q = q.filter(ProjectChangeOrder.status == status)
    cos = q.order_by(ProjectChangeOrder.created_at.desc()).all()
    phases = {p.id: p.name for p in db.query(Phase).filter(Phase.project_id == project_id).all()}
    approved, pending, n_approved = co_totals(db, project_id)
    budget = cf(project.budget or 0)
    today = date.today().isoformat()
    slug = "".join(ch if ch.isalnum() else "-" for ch in project.name.lower()).strip("-")

    def eff_cost(c):
        return cf(c.approved_cost if c.status == "Approved" and c.approved_cost is not None else c.estimated_cost)

    summary = [("Original Contract (Baseline)", budget),
               ("Approved Variations", approved),
               ("Revised Contract Value", round(budget + approved, 2)),
               ("Pending Review (not committed)", pending)]
    by_cat, by_phase = {}, {}
    for c in cos:
        if c.status == "Approved":
            by_cat[c.category] = by_cat.get(c.category, 0) + eff_cost(c)
            pname = phases.get(c.phase_id, "Unassigned")
            by_phase[pname] = by_phase.get(pname, 0) + eff_cost(c)
    filters_note = ", ".join(x for x in [
        f"Phase: {phases.get(phase_id, phase_id)}" if phase_id else "",
        f"Category: {category}" if category else "",
        f"Status: {status}" if status else ""] if x) or "All change orders"
    headers = ["CO #", "Title", "Phase", "Category", "Status", "Requested", "Est. Cost", "Approved Cost", "Days"]

    def rowvals(c):
        return [c.co_number, c.title, phases.get(c.phase_id, "-"), c.category, c.status,
                c.date_requested.isoformat() if c.date_requested else "-",
                cf(c.estimated_cost), cf(c.approved_cost) if c.approved_cost is not None else None,
                c.estimated_time_impact_days or 0]

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws["A1"] = f"SITERA — Change Orders / Variations: {project.name}"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = f"Generated {today} · Filter: {filters_note}"
        for i, (k, v) in enumerate(summary, start=4):
            ws.cell(row=i, column=1, value=k).font = BOLD
            ws.cell(row=i, column=2, value=v).number_format = "#,##0.00"
        r0 = 9
        ws.cell(row=r0, column=1, value="Approved Variations by Category").font = BOLD
        for i, (k, v) in enumerate(sorted(by_cat.items()), start=r0 + 1):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=v).number_format = "#,##0.00"
        r1 = r0 + len(by_cat) + 2
        ws.cell(row=r1, column=1, value="Approved Variations by Phase").font = BOLD
        for i, (k, v) in enumerate(sorted(by_phase.items()), start=r1 + 1):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=v).number_format = "#,##0.00"
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 18
        ws2 = wb.create_sheet("Change Orders")
        ws_header_row(ws2, 1, headers, [13, 36, 18, 18, 20, 12, 14, 14, 7])
        for j, c in enumerate(cos, start=2):
            for col, v in enumerate(rowvals(c), start=1):
                cell = ws2.cell(row=j, column=col, value=v)
                if col in (7, 8) and v is not None:
                    cell.number_format = "#,##0.00"
        return xlsx_response(wb, f"{slug}-change-orders-{today}.xlsx")

    st = _styles()
    story = [Paragraph(f"SITERA — Change Orders / Variations: {project.name}", st["title"]),
             Paragraph(f"Generated {today} · Filter: {filters_note}", st["sub"])]
    story.append(styled_table([[k for k, _ in summary], [money(v) for _, v in summary]],
                              col_widths=[45 * mm] * 4, left_cols=0))
    if by_cat:
        story.append(Paragraph("Approved Variations by Category", st["h2"]))
        story.append(styled_table([["Category", "Amount"]] + [[k, money(v)] for k, v in sorted(by_cat.items())],
                                  col_widths=[110 * mm, 40 * mm]))
    if by_phase:
        story.append(Paragraph("Approved Variations by Phase", st["h2"]))
        story.append(styled_table([["Phase", "Amount"]] + [[k, money(v)] for k, v in sorted(by_phase.items())],
                                  col_widths=[110 * mm, 40 * mm]))
    story.append(Paragraph("Change Order Register", st["h2"]))
    trows = [headers]
    for c in cos:
        v = rowvals(c)
        trows.append([v[0], Paragraph(v[1], st["cell"]), Paragraph(v[2], st["cell"]),
                      Paragraph(v[3], st["cell"]), Paragraph(v[4], st["cell"]), v[5],
                      money(v[6]), money(v[7]) if v[7] is not None else "-", str(v[8])])
    if len(trows) == 1:
        trows.append(["-"] * 9)
    story.append(styled_table(trows, col_widths=[16 * mm, 34 * mm, 22 * mm, 24 * mm, 24 * mm,
                                                 18 * mm, 21 * mm, 21 * mm, 10 * mm], left_cols=6))
    return pdf_response(story, f"{slug}-change-orders-{today}.pdf")
